import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image
from openpyxl.worksheet.header_footer import HeaderFooter

def cargar_datos_asistencia(ruta_json):
    with open(ruta_json, 'r', encoding='utf-8') as archivo:
        datos = json.load(archivo)
    return datos.get("estudiantes", [])

def crear_reporte_asistencia(nombre_archivo, ruta_json, ruta_imagen):
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    for col in range(3, 20):
        ws.column_dimensions[chr(64 + col)].width = 4
    ws.column_dimensions['T'].width = 8

    bold_font = Font(bold=True, size=13)
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Insertar imagen centrada en el área A1:B3
    img = Image(ruta_imagen)
    img.width = 100
    img.height = 120
    ws.add_image(img, 'A1')

    # Encabezados principales
    encabezados_principales = [
        ("C1:T1", "UNIVERSIDAD NACIONAL DE LA AMAZONIA PERUANA"),
        ("C2:T2", "ESCUELA DE POSTGRADO"),
        ("C3:T3", "MAESTRIA EN EDUCACIÓN CON MENCION EN GESTION EDUCATIVA-2024-I")
    ]

    for rango, texto in encabezados_principales:
        ws.merge_cells(rango)
        cell = ws[rango.split(":")[0]]
        cell.value = texto
        cell.font = bold_font
        cell.alignment = center_alignment

    datos_generales = [
        ("B4", "ASIGNATURA:", "C4:T4", "MDU-101: FILOSOFIA DE LA EDUCACION UNIVERSITARIA"),
        ("B5", "DOCENTE:", "C5:T5", "Dra. JULIANA SANCHEZ BABILONIA"),
        ("B6", "FECHA:", "C6:T6", "Del 08 al 30 de junio del 2024")
    ]

    for celdas1, texto1, celdas2, texto2 in datos_generales:
        ws.merge_cells(celdas1)
        ws.merge_cells(celdas2)
        ws[celdas1.split(":")[0]].value = texto1
        ws[celdas2.split(":")[0]].value = texto2
        ws[celdas1.split(":")[0]].font = header_font
        ws[celdas2.split(":")[0]].font = normal_font
        ws[celdas1.split(":")[0]].alignment = left_alignment
        ws[celdas2.split(":")[0]].alignment = left_alignment

    ws.merge_cells('C7:T7')
    ws['C7'] = "Control de Asistencia"
    ws['C7'].alignment = center_alignment
    ws['C7'].font = bold_font

    ws['B8'] = "APELLIDOS Y NOMBRES / secciones"
    ws['B8'].alignment = center_alignment
    ws['B8'].font = bold_font
    ws['B8'].border = thin_border

    for i in range(1, 18):
        cell = ws.cell(row=8, column=2 + i)
        cell.value = i
        cell.alignment = center_alignment
        cell.border = thin_border

    ws['T8'] = "TOTAL"
    ws['T8'].alignment = center_alignment
    ws['T8'].font = bold_font
    ws['T8'].border = thin_border

    estudiantes = cargar_datos_asistencia(ruta_json)

    for fila, estudiante in enumerate(estudiantes, start=9):
        ws.cell(row=fila, column=1, value=f"'{fila - 8:02}").alignment = center_alignment
        ws.cell(row=fila, column=1).border = thin_border

        ws.cell(row=fila, column=2, value=estudiante["nombre"]).border = thin_border

        total_asistencias = 0
        for j, asistencia in enumerate(estudiante["asistencia"], start=1):
            cell = ws.cell(row=fila, column=2 + j)
            cell.value = asistencia
            cell.alignment = center_alignment
            cell.border = thin_border
            if asistencia == "P":
                total_asistencias += 1

        ws.cell(row=fila, column=20, value=total_asistencias).alignment = center_alignment
        ws.cell(row=fila, column=20).border = thin_border

    # Pie de página visible completo
    ws.oddFooter.right.text = "ESCUELA DE POSTGRADO DE LA UNAP - Página &P"

    wb.save(nombre_archivo)


crear_reporte_asistencia("Asistencia.xlsx", "datos_asistencia.json", "C:/Users/Patrick/Desktop/programacion/exel/20.png")

#python exel.py